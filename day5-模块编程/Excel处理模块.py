# -*- coding: utf-8 -*- 
# Created By CHT  
"""
编码     : encoding="utf-8"
Time    : 2023/7/19 17:37
File    : Excel处理模块.py
Project : PycharmCode
"""
from openpyxl import Workbook, load_workbook

# ex1 = Workbook()  # 在内存创建一个Excel
# sheet = ex1.active  # 获取当前页sheet
# print(sheet.title)  # 打印当前页名
# sheet.title = "无敌是多磨寂寞"  # 修改当前页名
#
# sheet["B1"] = "girl"  # 写入数据
# sheet["B2"] = "cy"
# sheet.append(["girl1", "cy1"])  # 在已有数据的下一行添加数据
# ex1.save("test.xlsx")  # 保存文件
ex2 = load_workbook("test.xlsx")
print(ex2.sheetnames)  # 获取所有页名
sheet = ex2["无敌是多磨寂寞"]
print(sheet["B2"])
print(sheet["B2"].value)
